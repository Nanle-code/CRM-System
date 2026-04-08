#!/usr/bin/env python3
"""
Mini CRM System - Flask REST API Backend
Acts as a bridge between Excel file and frontend

GETTING STARTED (3 steps):
1. pip install -r requirements.txt
2. python setup_excel.py  # Generate customers.xlsx with sample data
3. python server.py       # Start the Flask server on http://localhost:5000

Then open index.html in your browser to use the CRM system.
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import uuid
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "customers.xlsx"

def get_customers_sheet():
    """Get the customers worksheet from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return None
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    return wb["Customers"]

def get_activity_log_sheet():
    """Get the activity log worksheet from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return None
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    return wb["Activity Log"]

def save_workbook(wb):
    """Save workbook to Excel file"""
    wb.save(EXCEL_FILE)

def row_to_dict(sheet, row_num, headers):
    """Convert a worksheet row to dictionary"""
    wb = sheet.parent
    row_data = {}
    for col_num, header in enumerate(headers, 1):
        cell_value = sheet.cell(row=row_num, column=col_num).value
        row_data[header] = cell_value
    return row_data

def get_headers(sheet):
    """Get headers from worksheet"""
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers

# API Routes

@app.route('/customers', methods=['GET'])
def get_all_customers():
    """Get all customers as JSON"""
    sheet = get_customers_sheet()
    if sheet is None:
        return jsonify({"error": "Excel file not found"}), 404
    
    headers = get_headers(sheet)
    customers = []
    
    for row_num in range(2, sheet.max_row + 1):
        customer = row_to_dict(sheet, row_num, headers)
        customers.append(customer)
    
    return jsonify(customers)

@app.route('/customers/<customer_id>', methods=['GET'])
def get_customer(customer_id):
    """Get a single customer with their activity log"""
    # Get customer
    customers_sheet = get_customers_sheet()
    if customers_sheet is None:
        return jsonify({"error": "Excel file not found"}), 404
    
    headers = get_headers(customers_sheet)
    customer = None
    
    for row_num in range(2, customers_sheet.max_row + 1):
        row_customer = row_to_dict(customers_sheet, row_num, headers)
        if row_customer.get('ID') == customer_id:
            customer = row_customer
            break
    
    if not customer:
        return jsonify({"error": "Customer not found"}), 404
    
    # Get activity log for this customer
    activity_sheet = get_activity_log_sheet()
    activity_log = []
    
    if activity_sheet:
        activity_headers = get_headers(activity_sheet)
        for row_num in range(2, activity_sheet.max_row + 1):
            activity = row_to_dict(activity_sheet, row_num, activity_headers)
            if activity.get('Customer ID') == customer_id:
                activity_log.append(activity)
    
    # Sort activity log by timestamp (newest first)
    activity_log.sort(key=lambda x: x.get('Timestamp', ''), reverse=True)
    
    return jsonify({
        "customer": customer,
        "activity_log": activity_log
    })

@app.route('/customers', methods=['POST'])
def add_customer():
    """Add a new customer"""
    data = request.get_json()
    
    # Validate required fields
    required_fields = ['Full Name', 'Email', 'Company']
    for field in required_fields:
        if field not in data or not data[field]:
            return jsonify({"error": f"Missing required field: {field}"}), 400
    
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    customers_sheet = wb["Customers"]
    
    # Generate new customer data
    new_customer = {
        'ID': str(uuid.uuid4()),
        'Full Name': data['Full Name'],
        'Email': data['Email'],
        'Phone': data.get('Phone', ''),
        'Company': data['Company'],
        'Status': data.get('Status', 'Lead'),
        'Deal Value': data.get('Deal Value', 0),
        'Tags': data.get('Tags', ''),
        'Created Date': datetime.now().strftime('%Y-%m-%d'),
        'Last Updated': datetime.now().strftime('%Y-%m-%d')
    }
    
    # Add new row
    headers = get_headers(customers_sheet)
    row_data = [new_customer.get(header, '') for header in headers]
    customers_sheet.append(row_data)
    
    # Save workbook
    save_workbook(wb)
    
    return jsonify(new_customer), 201

@app.route('/customers/<customer_id>', methods=['PUT'])
def update_customer(customer_id):
    """Update an existing customer"""
    data = request.get_json()
    
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    customers_sheet = wb["Customers"]
    
    headers = get_headers(customers_sheet)
    customer_found = False
    
    # Find and update customer
    for row_num in range(2, customers_sheet.max_row + 1):
        current_id = customers_sheet.cell(row=row_num, column=1).value
        if current_id == customer_id:
            customer_found = True
            # Update fields
            for col_num, header in enumerate(headers, 1):
                if header in data and header != 'ID':  # Don't update ID
                    customers_sheet.cell(row=row_num, column=col_num).value = data[header]
            
            # Update Last Updated timestamp
            last_updated_col = headers.index('Last Updated') + 1
            customers_sheet.cell(row=row_num, column=last_updated_col).value = datetime.now().strftime('%Y-%m-%d')
            break
    
    if not customer_found:
        return jsonify({"error": "Customer not found"}), 404
    
    # Save workbook
    save_workbook(wb)
    
    # Return updated customer
    updated_customer = row_to_dict(customers_sheet, row_num, headers)
    return jsonify(updated_customer)

@app.route('/customers/<customer_id>', methods=['DELETE'])
def delete_customer(customer_id):
    """Delete a customer"""
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE)
    customers_sheet = wb["Customers"]
    
    headers = get_headers(customers_sheet)
    customer_found = False
    
    # Find and delete customer
    for row_num in range(2, customers_sheet.max_row + 1):
        current_id = customers_sheet.cell(row=row_num, column=1).value
        if current_id == customer_id:
            customer_found = True
            # Delete the row
            customers_sheet.delete_rows(row_num)
            break
    
    if not customer_found:
        return jsonify({"error": "Customer not found"}), 404
    
    # Also delete related activity log entries
    activity_sheet = wb["Activity Log"]
    rows_to_delete = []
    
    for row_num in range(2, activity_sheet.max_row + 1):
        customer_id_col = activity_sheet.cell(row=row_num, column=2).value
        if customer_id_col == customer_id:
            rows_to_delete.append(row_num)
    
    # Delete activity log entries (from bottom to top to maintain row numbers)
    for row_num in reversed(rows_to_delete):
        activity_sheet.delete_rows(row_num)
    
    # Save workbook
    save_workbook(wb)
    
    return jsonify({"message": "Customer deleted successfully"})

@app.route('/customers/<customer_id>/log', methods=['POST'])
def add_activity_log(customer_id):
    """Add a new activity log entry for a customer"""
    data = request.get_json()
    
    # Validate required fields
    if 'Action' not in data or not data['Action']:
        return jsonify({"error": "Missing required field: Action"}), 400
    
    # Check if customer exists
    wb = openpyxl.load_workbook(EXCEL_FILE)
    customers_sheet = wb["Customers"]
    
    customer_found = False
    for row_num in range(2, customers_sheet.max_row + 1):
        current_id = customers_sheet.cell(row=row_num, column=1).value
        if current_id == customer_id:
            customer_found = True
            break
    
    if not customer_found:
        return jsonify({"error": "Customer not found"}), 404
    
    # Add activity log entry
    activity_sheet = wb["Activity Log"]
    
    new_log = {
        'Log ID': str(uuid.uuid4()),
        'Customer ID': customer_id,
        'Action': data['Action'],
        'Notes': data.get('Notes', ''),
        'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    headers = get_headers(activity_sheet)
    row_data = [new_log.get(header, '') for header in headers]
    activity_sheet.append(row_data)
    
    # Save workbook
    save_workbook(wb)
    
    return jsonify(new_log), 201

@app.route('/stats', methods=['GET'])
def get_stats():
    """Get aggregate statistics"""
    sheet = get_customers_sheet()
    if sheet is None:
        return jsonify({"error": "Excel file not found"}), 404
    
    headers = get_headers(sheet)
    customers = []
    
    for row_num in range(2, sheet.max_row + 1):
        customer = row_to_dict(sheet, row_num, headers)
        customers.append(customer)
    
    # Calculate stats
    total_customers = len(customers)
    
    # Count by status
    status_counts = {}
    total_deal_value = 0
    
    for customer in customers:
        status = customer.get('Status', 'Unknown')
        status_counts[status] = status_counts.get(status, 0) + 1
        
        # Add to total deal value
        deal_value = customer.get('Deal Value', 0)
        try:
            if isinstance(deal_value, str):
                deal_value = float(deal_value.replace(',', '').replace('$', ''))
            total_deal_value += deal_value
        except:
            pass
    
    # Count leads this month
    current_month = datetime.now().strftime('%Y-%m')
    leads_this_month = 0
    
    for customer in customers:
        created_date = customer.get('Created Date', '')
        if created_date.startswith(current_month) and customer.get('Status') == 'Lead':
            leads_this_month += 1
    
    stats = {
        'total_customers': total_customers,
        'status_counts': status_counts,
        'total_deal_value': total_deal_value,
        'leads_this_month': leads_this_month
    }
    
    return jsonify(stats)

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

if __name__ == '__main__':
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: {EXCEL_FILE} not found!")
        print("Please run 'python setup_excel.py' first to create the Excel file.")
    else:
        print("Starting Mini CRM Server...")
        print("API available at: http://localhost:5000")
        print("Frontend should be served from index.html")
        app.run(debug=True, host='0.0.0.0', port=5000)
