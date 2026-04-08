#!/usr/bin/env python3
"""
Setup Excel Backend for Mini CRM System
Creates customers.xlsx with sample data and styling
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import uuid

def create_customers_sheet(wb):
    """Create and populate the Customers sheet"""
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ["ID", "Full Name", "Email", "Phone", "Company", "Status", "Deal Value", "Tags", "Created Date", "Last Updated"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data
    sample_customers = [
        ["John Smith", "john.smith@techcorp.com", "+1-555-0123", "TechCorp Inc.", "Active", 75000, "enterprise,software", "2024-01-15", "2024-04-01"],
        ["Sarah Johnson", "sarah.j@marketing.com", "+1-555-0124", "Marketing Solutions", "Active", 45000, "marketing,retention", "2024-01-20", "2024-03-28"],
        ["Michael Chen", "m.chen@startup.io", "+1-555-0125", "StartupIO", "Lead", 120000, "startup,tech", "2024-02-01", "2024-04-05"],
        ["Emily Davis", "emily@retail.com", "+1-555-0126", "Retail Dynamics", "Churned", 25000, "retail,lost", "2024-01-10", "2024-02-15"],
        ["Robert Wilson", "rwilson@finance.com", "+1-555-0127", "Finance Plus", "Prospect", 95000, "finance,enterprise", "2024-02-15", "2024-04-02"],
        ["Lisa Anderson", "lisa@healthcare.org", "+1-555-0128", "Healthcare Systems", "Active", 60000, "healthcare,nonprofit", "2024-01-25", "2024-03-30"],
        ["David Martinez", "d.martinez@logistics.com", "+1-555-0129", "Global Logistics", "Lead", 85000, "logistics,global", "2024-02-10", "2024-04-04"],
        ["Jennifer Taylor", "jtaylor@education.edu", "+1-555-0130", "EduTech Solutions", "Active", 35000, "education,edtech", "2024-01-18", "2024-03-25"],
        ["James Brown", "james@consulting.com", "+1-555-0131", "Consulting Group", "Prospect", 110000, "consulting,b2b", "2024-02-20", "2024-04-06"],
        ["Maria Garcia", "m.garcia@restaurant.com", "+1-555-0132", "Restaurant Chain", "Churned", 30000, "restaurant,food", "2024-01-05", "2024-02-10"],
        ["Thomas Lee", "tlee@manufacturing.com", "+1-555-0133", "Manufacturing Co", "Active", 150000, "manufacturing,industrial", "2024-01-12", "2024-04-03"],
        ["Patricia White", "pwhite@legal.com", "+1-555-0134", "Legal Services", "Lead", 55000, "legal,professional", "2024-02-12", "2024-04-07"],
        ["Christopher Hall", "chall@realestate.com", "+1-555-0135", "Real Estate Pro", "Prospect", 80000, "realestate,property", "2024-02-18", "2024-04-08"],
        ["Nancy Clark", "n.clark@nonprofit.org", "+1-555-0136", "NonProfit Network", "Active", 20000, "nonprofit,ngo", "2024-01-22", "2024-03-27"],
        ["Daniel Lewis", "d.lewis@automotive.com", "+1-555-0137", "Auto Solutions", "Lead", 130000, "automotive,tech", "2024-02-05", "2024-04-09"]
    ]
    
    # Add customer data with UUIDs
    for customer in sample_customers:
        customer_id = str(uuid.uuid4())
        ws.append([customer_id] + customer)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_activity_log_sheet(wb):
    """Create and populate the Activity Log sheet"""
    ws = wb.create_sheet("Activity Log")
    
    # Headers
    headers = ["Log ID", "Customer ID", "Action", "Notes", "Timestamp"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A7C7E", end_color="4A7C7E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample activity log entries
    sample_activities = [
        ["Initial contact made via cold call", "2024-01-16 10:30:00"],
        ["Scheduled product demo", "2024-01-17 14:15:00"],
        ["Sent proposal document", "2024-01-18 09:45:00"],
        ["Follow-up email sent", "2024-01-19 16:20:00"],
        ["Contract negotiation started", "2024-01-22 11:00:00"],
        ["Deal closed successfully", "2024-01-25 15:30:00"],
        ["Customer onboarding completed", "2024-01-28 13:45:00"],
        ["First check-in call", "2024-02-05 10:00:00"],
        ["Renewal discussion initiated", "2024-02-15 14:30:00"],
        ["Customer feedback collected", "2024-02-20 11:15:00"],
        ["Issue resolved - technical support", "2024-02-25 16:45:00"],
        ["Quarterly business review", "2024-03-01 13:00:00"],
        ["Upsell opportunity identified", "2024-03-10 10:30:00"],
        ["Customer training session", "2024-03-15 14:00:00"],
        ["Account review meeting", "2024-03-20 11:30:00"],
        ["New feature request received", "2024-03-25 15:15:00"],
        ["Customer satisfaction survey", "2024-03-28 09:00:00"],
        ["Partnership discussion", "2024-04-01 16:00:00"],
        ["Expansion opportunity", "2024-04-03 12:30:00"],
        ["Renewal contract sent", "2024-04-05 14:45:00"]
    ]
    
    # Get customer IDs from the first sheet
    customers_ws = wb["Customers"]
    customer_ids = []
    for row in range(2, min(17, customers_ws.max_row + 1)):  # Get first 15 customer IDs
        customer_ids.append(customers_ws.cell(row=row, column=1).value)
    
    # Add activity log entries
    actions = ["Email", "Call", "Meeting", "Demo", "Follow-up", "Support", "Review", "Proposal"]
    
    for i, (notes, timestamp) in enumerate(sample_activities):
        log_id = str(uuid.uuid4())
        customer_id = customer_ids[i % len(customer_ids)]  # Cycle through customers
        action = actions[i % len(actions)]
        ws.append([log_id, customer_id, action, notes, timestamp])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Main function to create the Excel file"""
    print("Creating customers.xlsx with sample data...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create sheets
    create_customers_sheet(wb)
    create_activity_log_sheet(wb)
    
    # Save the file
    wb.save("customers.xlsx")
    print("Excel file created successfully!")
    print(f"Customers: {wb['Customers'].max_row - 1} records")
    print(f"Activity Log: {wb['Activity Log'].max_row - 1} records")

if __name__ == "__main__":
    main()
